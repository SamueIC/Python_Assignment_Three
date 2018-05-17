from pathlib import Path, PurePosixPath
from abc import ABCMeta, abstractmethod
from csv import DictReader as CSVDictReader
from openpyxl import load_workbook, utils
from datetime import datetime, date
from Interpreter.validator import Validator


# James
class FileHandler:
    # James
    def __init__(self, file_name):
        self.filename = file_name
        self.file_type = None
        # self.valid = Validator()

    # James
    def file_exist(self):
        """
        Checks if the current file exists
        """
        if self.filename.exists():
            return True
        else:
            return False

    # Wesley
    def set_file_type(self):
        """
        Will get the file type and will create the
        corresponding solid class and set it to self.file_type
        """
        suffix = PurePosixPath(self.filename).suffix
        print("The File's extension is: " + "'" + suffix + "'")
        file_types = {
            '.csv': FileTypeCSV(),
            '.xlsx': FileTypeXLSX(),
            '.txt': FileTypeTXT()
        }
        self.file_type = file_types[suffix]

    def read(self):
        return self.file_type.read(self.filename)


# Wesley
class FileTypeAbstract(metaclass=ABCMeta):
    def __init__(self):
        self.validator = Validator()
    # Wesley
    @abstractmethod
    def read(self, filename):
        pass


# Wesley
class FileTypeXLSX(FileTypeAbstract):
    # Wesley
    def read(self, filename):
        """
        Return dictionary with key => value pairs
        :param filename is the file where the values exist
        """
        data = dict()
        empno = 0
        keys = []
        a_row = 0
        try:
            workbook = load_workbook(filename)
            first_sheet = workbook.sheetnames[0]
            worksheet = workbook[first_sheet]
            for row in worksheet.iter_rows():
                record = dict()
                row_num = 0
                for cell in row:
                    a_row = cell.row
                    if 1 == a_row:
                        keys.append(cell.value)
                    else:
                        valid = cell.value
                        if isinstance(cell.value, datetime):
                            valid = Validator.xlsx_date(cell.value)
                        record[keys[row_num]] = valid
                    row_num += 1
                if a_row > 1:
                    data[empno] = record
                empno += 1
            # print(data)
            result = Validator.save_dict(data)
            return result
        except PermissionError:
            print("Sorry, you don't have enough permissions to access this file")

# The above function contains a date object in the dictionary for each date,
# as the birthday is a date, may need to access the values stored in the date object when validating


# Wesley
class FileTypeCSV(FileTypeAbstract):
    # James
    def read(self, filename):
        """
        Return dictionary with key => value pairs
        :param filename is the file where the values exist
        >>> a=FileTypeCSV()
        >>> FileTypeCSV.read(a,"Saves/data.csv")
        """
        try:
            data = dict()
            empno = 0
            with open(filename) as f:
                reader = CSVDictReader(f)
                for row in reader:
                    record = dict()
                    for key in row:
                        record[key] = row.get(key)
                    data[empno] = record
                    empno += 1
                # print(data)
            # James' changes (13/03)
            result = Validator.save_dict(data)
            return result
        except TypeError:
            print("Error!!")


class FileTypeTXT(FileTypeAbstract):
    def read(self, filename):
        """
        Return dictionary with key => value pairs
        :param filename is the file where the values exist
        >>> FileTypeTXT.read("Saves/data.txt")
        """
        file = open(filename, 'r')
        data = dict()
        empno = 0
        try:

            for line in file:  # FIRST LOOP
                rows = line.split(";")
                dictionary = dict()
                for row in rows:
                    if len(row.split("=")) == 2:
                        key = row.split("=")[0]
                        value = row.split("=")[1]
                        value = value.rstrip('\n')
                        dictionary[key] = value
                        data[empno] = dictionary
                    else:
                        print("File error")
                        raise ValueError
                empno += 1
            result = self.validator.save_dict(data)
            return result

        except Exception as e:
            print(e)
